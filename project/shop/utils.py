from io import BytesIO
from openpyxl import Workbook
from datetime import datetime

def generate_receipt(user, items, total_price):
    wb = Workbook()
    ws = wb.active
    ws.title = "Чек"
    
    ws['A1'] = "Чек по заказу"
    ws['A2'] = f"Покупатель: {user.username}"
    ws['A3'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    
    headers = ['Товар', 'Количество', 'Цена за шт.', 'Сумма']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)
    
    current_row = 6
    for item in items:
        item_total = item.product.price * item.quantity
        
        ws.cell(row=current_row, column=1, value=item.product.name)
        ws.cell(row=current_row, column=2, value=item.quantity)
        ws.cell(row=current_row, column=3, value=float(item.product.price))
        ws.cell(row=current_row, column=4, value=float(item_total))
        
        current_row += 1
    
    ws.cell(row=current_row + 1, column=3, value="ИТОГО:")
    ws.cell(row=current_row + 1, column=4, value=float(total_price))
    
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream